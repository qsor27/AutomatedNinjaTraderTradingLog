import pandas as pd
import ast
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import shutil

def parse_exits(exits_str):
    try:
        exits = ast.literal_eval(exits_str)
    except:
        print(f"Failed to parse exits: {exits_str}")
        return []
    
    parsed_exits = []
    for exit in exits:
        exit_type, exit_data = exit.split(': ', 1)
        try:
            exit_dict = ast.literal_eval(exit_data)
            exit_dict['type'] = exit_type
            parsed_exits.append(exit_dict)
        except:
            print(f"Failed to parse exit data: {exit_data}")
    
    return parsed_exits

def parse_date(date_string):
    date_formats = [
        '%Y-%m-%d %H:%M:%S.%f',
        '%m/%d/%Y %I:%M:%S %p',
        '%Y-%m-%d %H:%M:%S',
        '%m/%d/%Y %H:%M:%S',
        '%m/%d/%Y %H:%M',  # Added this format to handle the current date string
    ]
    
    for date_format in date_formats:
        try:
            return datetime.strptime(date_string, date_format).date()
        except ValueError:
            continue
    
    raise ValueError(f"Unable to parse date: {date_string}")

def process_csv(input_csv):
    df = pd.read_csv(input_csv)
    
    results = []
    for _, row in df.iterrows():
        exits = parse_exits(row['Exits'])
        
        try:
            entry_date = parse_date(row['Entry_time'])
        except ValueError as e:
            print(f"Error parsing date: {e}")
            continue
        
        result = {
            'Account': row['Account'],
            'Date': entry_date,
            'Instrument': row['Instrument'],
            'Entry Time': row['Entry_time'],
            'Entry Price': row['Entry_price'],
            'Market Position': row['Market_pos'],
            'Qty': row['Qty'],
            'Commission': row['Total_commission'],
            'Stop Loss': None,
            'TP 1': None,
            'TP 1 Qty': None,
            'TP 1 Exit Time': None,
            'TP 2': None,
            'TP 2 Qty': None,
            'TP 2 Exit Time': None,
            'TP 3': None,
            'TP 3 Qty': None,
            'TP 3 Exit Time': None,
            'SL Execution': None,
            'SL Exit Time': None,
            'RoE Pts': 0,
            'RoE $': 0
        }
        
        tp_count = 0
        sl_triggered = False
        for exit in exits:
            if exit['type'].startswith('TP'):
                tp_count += 1
                result[f'TP {tp_count}'] = exit['price']
                result[f'TP {tp_count} Qty'] = exit['Qty']
                result[f'TP {tp_count} Exit Time'] = exit['Exit_time']
                result['RoE Pts'] += exit['Pnl_points']
                result['RoE $'] += exit['Pnl_dollars']
            elif exit['type'].startswith('SL'):
                result['Stop Loss'] = exit['price']
                result['SL Execution'] = exit['price']
                result['SL Exit Time'] = exit['Exit_time']
                result['RoE Pts'] = -abs(exit['Pnl_points'])
                result['RoE $'] = -abs(exit['Pnl_dollars'])
                sl_triggered = True
        
        if sl_triggered:
            result['RoE Pts'] = -abs(result['RoE Pts'])
            result['RoE $'] = -abs(result['RoE $'])
        
        results.append(result)
    
    return pd.DataFrame(results)

def update_excel(excel_file, sheet_name, df):
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
    else:
        sheet = workbook[sheet_name]
    
    # Find the first empty row
    first_empty_row = sheet.max_row + 1
    
    # If the sheet is empty, add headers
    if first_empty_row == 1:
        headers = list(df.columns)
        sheet.append(headers)
        first_empty_row = 2
    
    # Append new data
    for _, row in df.iterrows():
        sheet.append(row.tolist())
    
    workbook.save(excel_file)

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    archive_dir = os.path.join(script_dir, 'archive')
    excel_file = "TradingJournal.xlsx"
    sheet_name = "CurrentMonth"
    
    # Create archive directory if it doesn't exist
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)
    
    # Process all CSV files in the script directory
    for filename in os.listdir(script_dir):
        if filename.endswith('.csv'):
            input_csv = os.path.join(script_dir, filename)
            try:
                processed_df = process_csv(input_csv)
                update_excel(excel_file, sheet_name, processed_df)
                print(f"Data from {filename} has been appended to {excel_file} in the {sheet_name} sheet.")
                
                # Move processed file to archive
                shutil.move(input_csv, os.path.join(archive_dir, filename))
                print(f"Moved {filename} to archive.")
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")

if __name__ == "__main__":
    main()